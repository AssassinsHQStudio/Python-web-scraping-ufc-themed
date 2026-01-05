import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from datetime import datetime

UFC_event_number = 324
output_file = f"ufc_{UFC_event_number}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

def clean_fighter_name(name):
    """Clean fighter name by removing unwanted keywords and prefixes"""
    # Remove extra whitespace
    name = re.sub(r'\s+', ' ', name.strip())
    
    # Remove ranking numbers if present (e.g., "#4 Justin Gaethje" -> "Justin Gaethje")
    name = re.sub(r'^#\d+\s+', '', name)
    
    # Remove "C " prefix (champion indicator)
    name = re.sub(r'^C\s+', '', name)
    
    # Keywords to remove from the beginning
    prefix_keywords = ['Live now', 'Live', 'LIVE NOW', 'LIVE', 'Card', 'Method', 'Main Card', 'Main']
    for keyword in prefix_keywords:
        # Remove keyword at the start (case insensitive)
        name = re.sub(r'^' + re.escape(keyword) + r'\s+', '', name, flags=re.I)
    
    # Keywords to remove from the end
    suffix_keywords = ['Round', 'Round Time', 'Time', 'Follow live', 'Follow', 'LIVE NOW', 'LIVE', 'now']
    for keyword in suffix_keywords:
        # Remove keyword at the end (case insensitive)
        name = re.sub(r'\s+' + re.escape(keyword) + r'$', '', name, flags=re.I)
    
    # Clean up any remaining extra whitespace
    name = re.sub(r'\s+', ' ', name.strip())
    
    return name

def normalize_name_for_matching(name):
    """Normalize name for comparison - get last name and optionally first name"""
    # Split into words and get the last word (surname)
    words = name.split()
    if len(words) >= 1:
        return words[-1].lower()  # Return last name in lowercase
    return name.lower()

def names_match(name1, name2):
    """Check if two names refer to the same fighter"""
    name1_lower = name1.lower()
    name2_lower = name2.lower()
    
    # Exact match
    if name1_lower == name2_lower:
        return True
    
    # One name contains the other (e.g., "Gaethje" in "Justin Gaethje")
    if name1_lower in name2_lower or name2_lower in name1_lower:
        # But make sure it's not just a partial word match
        # Check if the shorter name is the last name of the longer name
        shorter = name1 if len(name1) < len(name2) else name2
        longer = name2 if len(name1) < len(name2) else name1
        
        shorter_words = shorter.lower().split()
        longer_words = longer.lower().split()
        
        # If shorter name's last word matches longer name's last word
        if shorter_words and longer_words and shorter_words[-1] == longer_words[-1]:
            return True
    
    return False

def get_fighter_pairs_from_ufc_event(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')

    # Extract text content and find "Name vs Name" patterns
    text_content = soup.get_text(separator=" ")
    
    # Pattern to match fighter names with "vs" between them
    # Handles names with multiple words, hyphens, apostrophes, and rankings
    pattern = re.compile(r'\b([A-Z][a-zA-Z\s\-\']{2,50})\s+vs\s+([A-Z][a-zA-Z\s\-\']{2,50})\b')
    matches = pattern.findall(text_content)
    
    # Use dictionary to store best version of each pair (prefer cleaner, longer names)
    seen_pairs = {}
    
    for match in matches:
        fighter1_raw = match[0].strip()
        fighter2_raw = match[1].strip()
        
        # Clean the names
        fighter1 = clean_fighter_name(fighter1_raw)
        fighter2 = clean_fighter_name(fighter2_raw)
        
        # Filter out invalid names
        exclude_patterns = [
            r'^(vs|odds|Flag)$',
            r'^(United States|England|Brazil|China|Russia|Dominican Republic|Lithuania|Cameroon)$',
            r'.*(Title|Bout|Interim|Women|Lightweight|Bantamweight|Heavyweight|Featherweight|Light Heavyweight|Middleweight|Flyweight|Fight Card).*',
        ]
        
        # Check if names are valid (must be proper names, not too short/long)
        valid1 = (len(fighter1) >= 5 and len(fighter1) <= 40 and 
                 not any(re.match(pattern, fighter1, re.I) for pattern in exclude_patterns) and
                 re.match(r'^[A-Z][a-zA-Z\s\-\']+$', fighter1))
        
        valid2 = (len(fighter2) >= 5 and len(fighter2) <= 40 and 
                 not any(re.match(pattern, fighter2, re.I) for pattern in exclude_patterns) and
                 re.match(r'^[A-Z][a-zA-Z\s\-\']+$', fighter2))
        
        if valid1 and valid2:
            # Check if this pair matches any existing pair (handles partial name matches)
            matched_key = None
            for existing_key, existing_pair in seen_pairs.items():
                existing_f1, existing_f2 = existing_pair
                # Check if fighters match (handles "Gaethje" vs "Justin Gaethje" cases)
                if (names_match(fighter1, existing_f1) and names_match(fighter2, existing_f2)) or \
                   (names_match(fighter1, existing_f2) and names_match(fighter2, existing_f1)):
                    matched_key = existing_key
                    break
            
            if matched_key is None:
                # New pair, add it with normalized key
                pair_key = tuple(sorted([normalize_name_for_matching(fighter1), normalize_name_for_matching(fighter2)]))
                seen_pairs[pair_key] = (fighter1, fighter2)
            else:
                # This pair matches an existing one, prefer the better version
                existing = seen_pairs[matched_key]
                current_total_len = len(fighter1) + len(fighter2)
                existing_total_len = len(existing[0]) + len(existing[1])
                
                # Prefer versions with full names (more words = more complete)
                current_words = len(fighter1.split()) + len(fighter2.split())
                existing_words = len(existing[0].split()) + len(existing[1].split())
                
                # Also prefer versions without common unwanted words
                current_has_unwanted = any(word in fighter1.lower() or word in fighter2.lower() 
                                          for word in ['live', 'round', 'method', 'card', 'follow', 'time'])
                existing_has_unwanted = any(word in existing[0].lower() or word in existing[1].lower() 
                                           for word in ['live', 'round', 'method', 'card', 'follow', 'time'])
                
                # Prefer cleaner version (no unwanted words), or more complete (more words), or longer
                if not current_has_unwanted and existing_has_unwanted:
                    seen_pairs[matched_key] = (fighter1, fighter2)
                elif current_has_unwanted == existing_has_unwanted:
                    if current_words > existing_words:
                        seen_pairs[matched_key] = (fighter1, fighter2)
                    elif current_words == existing_words and current_total_len > existing_total_len:
                        seen_pairs[matched_key] = (fighter1, fighter2)
    
    # Convert dictionary values to list
    fight_pairs = list(seen_pairs.values())
    
    return fight_pairs

def create_excel_file(fight_pairs, filename="ufc_fights.xlsx"):
    """Create an Excel file with fighter pairs"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fight Card"
    
    # Set headers
    ws['A1'] = "fighter 1"
    ws['B1'] = "fighter 2"
    
    # Make headers bold
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font
    
    # Add fighter pairs
    for row_idx, (fighter1, fighter2) in enumerate(fight_pairs, start=2):
        ws[f'A{row_idx}'] = fighter1
        ws[f'B{row_idx}'] = fighter2
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    
    # Save the file
    wb.save(filename)
    print(f"\nExcel file created: {filename}")
    return filename

# Example usage
event_url = F"https://www.ufc.com/event/ufc-{UFC_event_number}"
fight_pairs = get_fighter_pairs_from_ufc_event(event_url)
print("Fights found:")
for fighter1, fighter2 in fight_pairs:
    print(f"{fighter1} vs {fighter2}")

# Create Excel file
if fight_pairs:
    excel_filename = create_excel_file(fight_pairs, output_file)
    print(f"Total fights saved: {len(fight_pairs)}")
else:
    print("No fights found to save.")
